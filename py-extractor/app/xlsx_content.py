from __future__ import annotations

import io
from dataclasses import dataclass
from itertools import zip_longest

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

from app import markdown_table


@dataclass(frozen=True)
class XlsxContent:
    text: str
    table_count: int
    paragraph_count: int


def extract(data: bytes) -> XlsxContent:
    """Read workbook structure and cached values without evaluating formulas.

    The two read-only views are intentional: the formula view preserves the
    workbook's source expression, while the data-only view supplies a cached
    result when the file contains one. Pairing them by worksheet position
    keeps an uncached formula distinguishable from a genuinely empty cell.
    """
    formulas = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    cached = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sections: list[str] = []
    table_count = 0
    try:
        for formula_sheet, cached_sheet in zip(formulas.worksheets, cached.worksheets):
            rows: list[list[object]] = []
            array_formulas: dict[str, str] = {}
            for row_index, (formula_row, cached_row) in enumerate(
                zip(formula_sheet.iter_rows(), cached_sheet.iter_rows()), start=1
            ):
                values: list[object] = []
                for column_index, (formula_cell, cached_cell) in enumerate(
                    zip_longest(formula_row, cached_row), start=1
                ):
                    formula_value = getattr(formula_cell, "value", None)
                    cached_value = getattr(cached_cell, "value", None)
                    coordinate = f"{get_column_letter(column_index)}{row_index}"
                    values.append(
                        _format_cell(
                            formula_cell,
                            formula_value,
                            cached_value,
                            coordinate,
                            array_formulas,
                        )
                    )
                if any(str(cell).strip() for cell in values):
                    rows.append(values)
            if rows:
                sections.append(f"## {formula_sheet.title}\n\n{markdown_table.serialize(rows)}")
                table_count += 1
    finally:
        formulas.close()
        cached.close()
    return XlsxContent(
        text="\n\n".join(sections),
        table_count=table_count,
        paragraph_count=len(sections),
    )


def _format_cell(
    cell: object,
    source: object,
    cached: object,
    coordinate: str,
    array_formulas: dict[str, str],
) -> str:
    formula, array_ref = _formula_source(cell, source)
    if formula is None:
        return "" if cached is None else str(cached)

    if array_ref:
        if formula == "=":
            formula = array_formulas.get(array_ref, f"=array formula ({array_ref})")
        elif array_ref not in array_formulas:
            array_formulas[array_ref] = formula
        formula = f"{formula} [array cell {coordinate} of {array_ref}]"

    if cached is None:
        return f"{formula} [no cached value]"
    return f"{formula} [cached value: {cached}]"


def _formula_source(cell: object, value: object) -> tuple[str | None, str | None]:
    if isinstance(value, ArrayFormula):
        return value.text or "=", str(value.ref)
    if getattr(cell, "data_type", None) == "f":
        return str(value), None
    return None, None
